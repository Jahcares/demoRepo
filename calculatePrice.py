import openpyxl as xl

wb = xl.load_workbook('pythonTransaction.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1,1)

for row in range(2, sheet.max_row + 1):
    correctedPriceCellTitle = sheet.cell(1,4)
    correctedPriceCellTitle.value = 'CorrectPrice'
    cell = sheet.cell(row,3)
    correctedPrice = cell.value * 0.9
    correctedPriceCell = sheet.cell(row, 4)
    correctedPriceCell.value = correctedPrice

wb.save('transaction2.xlsx')